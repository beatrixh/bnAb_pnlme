from __future__ import annotations

import itertools
from pathlib import Path

import openpyxl

TRACKER_XLSX = Path("/mnt/c/Users/bhaddock/repos/bnAb_pnlme/run_plate_tagged_models/model_tracker.xlsx")

# Row layout in Sheet1, fixed by the existing m0/m1 columns.
ROWS = {
    "U_random": 3, "U_mab_virus": 4, "U_goes_down": 5,
    "L_random": 6, "L_mab_virus": 7, "L_goes_down": 8,
    "m_random": 9, "m_mab_virus": 10, "m_goes_down": 11,
    "e_random": 12, "e_mab_virus": 13, "e_goes_down": 14,
    "alpha_random": 15, "alpha_run_id": 16,
    "k_random": 17, "k_run_id": 18,
}

YES, NO = "Yes", "No"

# (random_effect, mab_virus_effect) states toggled per parameter; goes_down is
# always Yes and is not part of the factorial per the user's instruction.
STATES = [(NO, NO), (NO, YES), (YES, NO), (YES, YES)]


def main() -> None:
    wb = openpyxl.load_workbook(TRACKER_XLSX)
    ws = wb.active

    header = [c.value for c in ws[1]]
    existing_indices = [
        int(v[1:]) for v in header if isinstance(v, str) and v.startswith("m") and v[1:].isdigit()
    ]
    next_index = max(existing_indices, default=-1) + 1
    next_col = ws.max_column + 1

    combos = itertools.product(STATES, STATES, STATES)  # (L, m, e)
    for combo in combos:
        (l_random, l_mab), (m_random, m_mab), (e_random, e_mab) = combo

        col = next_col
        model_name = f"m{next_index}"

        ws.cell(row=1, column=col, value=model_name)

        ws.cell(row=ROWS["U_random"], column=col, value=NO)
        ws.cell(row=ROWS["U_mab_virus"], column=col, value=NO)
        ws.cell(row=ROWS["U_goes_down"], column=col, value=NO)

        ws.cell(row=ROWS["L_random"], column=col, value=l_random)
        ws.cell(row=ROWS["L_mab_virus"], column=col, value=l_mab)
        ws.cell(row=ROWS["L_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["m_random"], column=col, value=m_random)
        ws.cell(row=ROWS["m_mab_virus"], column=col, value=m_mab)
        ws.cell(row=ROWS["m_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["e_random"], column=col, value=e_random)
        ws.cell(row=ROWS["e_mab_virus"], column=col, value=e_mab)
        ws.cell(row=ROWS["e_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["alpha_random"], column=col, value=NO)
        ws.cell(row=ROWS["alpha_run_id"], column=col, value=YES)
        ws.cell(row=ROWS["k_random"], column=col, value=NO)
        ws.cell(row=ROWS["k_run_id"], column=col, value=YES)

        next_index += 1
        next_col += 1

    wb.save(TRACKER_XLSX)
    print(f"added {next_index - max(existing_indices, default=-1) - 1} columns, "
          f"through m{next_index - 1}")


if __name__ == "__main__":
    main()
