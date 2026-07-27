from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

MODELS_DIR = Path("/mnt/c/Users/bhaddock/repos/bnAb_pnlme/run_plate_tagged_models/model_files")
TRACKER_XLSX = Path("/mnt/c/Users/bhaddock/repos/bnAb_pnlme/run_plate_tagged_models/model_tracker.xlsx")

EFFECT_LABEL = {
    "Random effect": "random",
    "mab_virus fixed effect": "mab_virus",
    "goes_down fixed effect": "goes_down",
    "run_id fixed effect": "run_id",
}


def read_model_tracker(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    model_cols = {
        idx: name for idx, name in enumerate(header, start=1)
        if name and idx > 2
    }

    rows: dict[str, dict[str, str]] = {name: {} for name in model_cols.values()}
    current_param = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value:
            current_param = row[0].value
        effect = row[1].value
        if effect is None or current_param is None:
            continue
        toggle_name = f"{current_param}_{EFFECT_LABEL.get(effect, effect)}"
        for col_idx, model_name in model_cols.items():
            rows[model_name][toggle_name] = row[col_idx - 1].value

    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "model"
    return df.reset_index()


def read_criteria(model_name: str, models_dir: Path) -> dict:
    path = models_dir / model_name / "loglik.csv"
    if not path.exists():
        return {"model": model_name, "AIC": None, "BICc": None}
    ll = pd.read_csv(path)

    def first_matching(suffix: str):
        cols = [c for c in ll.columns if c.endswith(suffix)]
        return ll[cols[0]].iloc[0] if cols else None

    return {
        "model": model_name,
        "AIC": first_matching("AIC"),
        "BICc": first_matching("BICc"),
    }


def report_bicc_table(model_names: list[str], models_dir: Path, tracker_path: Path) -> pd.DataFrame:
    toggles = read_model_tracker(tracker_path)
    criteria = pd.DataFrame([read_criteria(m, models_dir) for m in model_names])

    missing = sorted(set(model_names) - set(toggles["model"]))
    if missing:
        print(f"warning: model(s) not found in tracker: {', '.join(missing)}")

    result = criteria.merge(toggles, on="model", how="left")
    toggle_cols = [c for c in toggles.columns if c != "model"]
    result = result[["model"] + toggle_cols + ["AIC", "BICc"]]
    return result.sort_values("BICc", na_position="last").reset_index(drop=True)


def main() -> None:
    # Edit this to whichever models you want reported.
    model_names = ["m0", "m1", "m2", "m3", "m4", "m5", "m6",
                    "m11", "m12", "m13", "m14", "m15", "m16", "m17"]

    result = report_bicc_table(model_names, MODELS_DIR, TRACKER_XLSX)
    print(result.to_string(index=False))
    result.to_csv(MODELS_DIR / "bicc_report.csv", index=False)


if __name__ == "__main__":
    main()
